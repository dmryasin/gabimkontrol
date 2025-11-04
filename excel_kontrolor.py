import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import os
import sys

APP_VERSION = "1.3.3"


class ExcelKontrolProgram:
    def __init__(self, root):
        self.root = root
        self.root.title("GABİM KONTROL PROGRAMI")
        self.root.geometry("800x600")
        self.root.resizable(False, False)
        # Uygulama başlığı (pencere başlığı)
        self.root.title("Gabim Kontrol")

        # Logo ikonlarını ayarla
        try:
            # PyInstaller ile paketlendiğinde dosya yolunu al
            if getattr(sys, 'frozen', False):
                base_path = sys._MEIPASS
            else:
                base_path = os.path.dirname(os.path.abspath(__file__))
            # Pencere küçük ikonu için ICO kullan
            logo_small_ico = os.path.join(base_path, 'logo_1.ico')
            if os.path.exists(logo_small_ico):
                try:
                    self.root.iconbitmap(default=logo_small_ico)
                except Exception:
                    pass
            # Pencere sol üst ve küçük ikonlar için: logo_1.png
            logo_small_path = os.path.join(base_path, 'logo_1.png')
            if os.path.exists(logo_small_path):
                self._icon_small = tk.PhotoImage(file=logo_small_path)
                # Pencere ve görev çubuğu için küçük icon ayarla (taskbar büyük ikon zaten EXE'den gelir)
                self.root.iconphoto(True, self._icon_small)
                self.root.wm_iconphoto(True, self._icon_small)
        except Exception as e:
            print(f"İkon yüklenemedi: {e}")

        self.excel_path = None
        self.workbook = None
        self.worksheet = None
        self.change_log = []  # Değişiklik logları

        self.setup_ui()

    def setup_ui(self):
        # Başlık
        title_frame = tk.Frame(self.root, bg="#2c3e50", height=80)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)

        # Logo ekle (sol tarafta)
        try:
            if getattr(sys, 'frozen', False):
                base_path = sys._MEIPASS
            else:
                base_path = os.path.dirname(os.path.abspath(__file__))

            logo_path = os.path.join(base_path, 'logo_1.png')
            if False and os.path.exists(logo_path):
                logo_image = tk.PhotoImage(file=logo_path)
                # Logo boyutunu ayarla (max 60x60 piksel)
                logo_label = tk.Label(title_frame, image=logo_image, bg="#2c3e50")
                logo_label.image = logo_image  # Referansı sakla
                logo_label.pack(side=tk.LEFT, padx=20, pady=10)
        except Exception as e:
            print(f"Logo yüklenemedi: {e}")

        title_label = tk.Label(
            title_frame,
            text=f"GABİM KONTROL PROGRAMI v{APP_VERSION}",
            font=("Arial", 18, "bold"),
            bg="#2c3e50",
            fg="white"
        )
        title_label.pack(side=tk.LEFT, pady=20)
        # Header metni: sadece program adı ve sürüm
        try:
            title_label.config(text=f"Gabim Kontrol v{APP_VERSION}")
        except Exception:
            pass

        # Ana içerik alanı
        main_frame = tk.Frame(self.root, bg="#ecf0f1")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # Dosya seçimi bölümü
        file_frame = tk.LabelFrame(
            main_frame,
            text="Dosya Seçimi",
            font=("Arial", 12, "bold"),
            bg="#ecf0f1",
            padx=10,
            pady=10
        )
        file_frame.pack(fill=tk.X, pady=(0, 20))

        self.file_label = tk.Label(
            file_frame,
            text="Henüz dosya seçilmedi",
            font=("Arial", 10),
            bg="#ecf0f1",
            fg="#7f8c8d"
        )
        self.file_label.pack(side=tk.LEFT, padx=(0, 10))

        select_btn = tk.Button(
            file_frame,
            text="Excel Dosyası Seç",
            command=self.select_file,
            bg="#3498db",
            fg="white",
            font=("Arial", 10, "bold"),
            padx=20,
            pady=5,
            cursor="hand2"
        )
        select_btn.pack(side=tk.RIGHT)

        # İşlem kuralları bölümü
        rules_frame = tk.LabelFrame(
            main_frame,
            text="Uygulanacak Kurallar",
            font=("Arial", 12, "bold"),
            bg="#ecf0f1",
            padx=10,
            pady=10
        )
        rules_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 20))

        rules_text = tk.Text(
            rules_frame,
            height=15,
            font=("Arial", 9),
            bg="white",
            wrap=tk.WORD,
            state=tk.DISABLED
        )
        rules_text.pack(fill=tk.BOTH, expand=True)

        rules_content = """
• C=3 ve AD=1 olan satırların AD ve A hücrelerini kırmızı yap
• AU > mevcut yıl ise AU hücrelerini kırmızı yap
• AU>2000 ve C=3 olan satırların AV değerini 7 yap
• C=15 olan satırların U değerini 3 yap
• C=3 ve CY<50 olan satırların BL değerini 2 yap
• C=3 ve CY<100 olan satırların BN değerini 2 yap
• C=3 ve CY<100 olan satırların BO değerini 1 yap
• Tüm satırlarda BP değeri max 3 olsun (fazla ise düzelt)
• BU > S ise BU'yu S değerine eşitle
• DB sütunundaki değerleri virgülden sonra max 2 haneye yuvarla
• C=1 veya C=2 ve W=2 ise AC ve AD değerleri 1 olmalı
• AU değeri 4 haneden fazla ya da az ise B ve AU hücrelerini yeşil yap
• C=3, 6 ya da 15 olan satırlarda AZ boş ise AZ=998 yap
• C=3 veya C=6 ise AW değerini 1 yap
• C=15 ise AW değerini 4 yap
• C=6 ve CS boş ise CS=2 yap
        """
        rules_text.config(state=tk.NORMAL)
        rules_text.insert("1.0", rules_content.strip())
        rules_text.config(state=tk.DISABLED)

        # İşlem butonları
        button_frame = tk.Frame(main_frame, bg="#ecf0f1")
        button_frame.pack(fill=tk.X)

        process_btn = tk.Button(
            button_frame,
            text="Kontrolleri Başlat",
            command=self.process_excel,
            bg="#27ae60",
            fg="white",
            font=("Arial", 12, "bold"),
            padx=30,
            pady=10,
            cursor="hand2"
        )
        process_btn.pack(side=tk.LEFT, padx=(0, 10))

        exit_btn = tk.Button(
            button_frame,
            text="Çıkış",
            command=self.root.quit,
            bg="#e74c3c",
            fg="white",
            font=("Arial", 12, "bold"),
            padx=30,
            pady=10,
            cursor="hand2"
        )
        exit_btn.pack(side=tk.RIGHT)

        # Durum çubuğu
        self.status_var = tk.StringVar()
        self.status_var.set("Hazır")
        status_bar = tk.Label(
            self.root,
            textvariable=self.status_var,
            bg="#34495e",
            fg="white",
            font=("Arial", 9),
            anchor=tk.W,
            padx=10
        )
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)

    def select_file(self):
        file_path = filedialog.askopenfilename(
            title="Excel Dosyası Seç",
            filetypes=[("Excel Dosyaları", "*.xlsx *.xls"), ("Tüm Dosyalar", "*.*")]
        )

        if file_path:
            self.excel_path = file_path
            file_name = os.path.basename(file_path)
            self.file_label.config(text=f"Seçilen dosya: {file_name}", fg="#2c3e50")
            self.status_var.set(f"Dosya seçildi: {file_name}")

    def get_cell_value(self, row, col_letter):
        """Hücre değerini güvenli şekilde al"""
        try:
            cell = self.worksheet[f"{col_letter}{row}"]
            value = cell.value
            if value is None:
                return None
            if isinstance(value, str):
                value = value.strip()
                if value == "":
                    return None
            return value
        except Exception:
            return None

    def log_change(self, row, col_letter, old_value, new_value, rule_description):
        """Değişikliği kaydet"""
        self.change_log.append({
            "satir": row,
            "sutun": col_letter,
            "eski_deger": old_value,
            "yeni_deger": new_value,
            "kural": rule_description
        })

    def set_cell_value(self, row, col_letter, value, rule_description=""):
        """Hücre değerini güvenli şekilde ayarla ve kaydet. True döndürürse değişiklik yapıldı."""
        try:
            old_value = self.get_cell_value(row, col_letter)

            # Eski değer ile yeni değeri karşılaştır
            old_num = self.to_number(old_value)
            new_num = self.to_number(value)

            # Eğer her ikisi de sayısal ise ve eşitse değişiklik yapma
            if old_num is not None and new_num is not None:
                if old_num == new_num:
                    return False  # Değişiklik yok
            # Eğer her ikisi de aynı string ise değişiklik yapma
            elif old_value == value:
                return False  # Değişiklik yok

            # Değişiklik varsa güncelle ve kaydet
            self.worksheet[f"{col_letter}{row}"] = value
            self.log_change(row, col_letter, old_value, value, rule_description)
            return True  # Değişiklik yapıldı
        except Exception as e:
            print(f"Hücre değeri ayarlanamadı {col_letter}{row}: {e}")
            return False

    def set_cell_red(self, row, col_letter, rule_description=""):
        """Hücreyi kırmızı yap ve kaydet. True döndürürse değişiklik yapıldı."""
        try:
            old_value = self.get_cell_value(row, col_letter)
            red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            self.worksheet[f"{col_letter}{row}"].fill = red_fill
            self.log_change(row, col_letter, old_value, f"{old_value} (KIRMIZI)", rule_description)
            return True
        except Exception as e:
            print(f"Hücre renklendirilemedi {col_letter}{row}: {e}")
            return False

    def set_cell_green(self, row, col_letter, rule_description=""):
        """Hücreyi yeşil yap ve kaydet. True döndürürse değişiklik yapıldı."""
        try:
            old_value = self.get_cell_value(row, col_letter)
            green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            self.worksheet[f"{col_letter}{row}"].fill = green_fill
            self.log_change(row, col_letter, old_value, f"{old_value} (YEŞİL)", rule_description)
            return True
        except Exception as e:
            print(f"Hücre renklendirilemedi {col_letter}{row}: {e}")
            return False

    def to_number(self, value):
        """Değeri sayıya çevir"""
        if value is None:
            return None
        try:
            if isinstance(value, (int, float)):
                return float(value)
            if isinstance(value, str):
                value = value.replace(",", ".")
                return float(value)
        except (ValueError, AttributeError):
            return None
        return None

    def save_change_log(self, file_path):
        """Değişiklik logunu dosyaya kaydet"""
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write("=" * 80 + "\n")
                f.write("EXCEL DEĞİŞİKLİK RAPORU\n")
                f.write("=" * 80 + "\n")
                f.write(f"Tarih: {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}\n")
                f.write(f"İşlenen Dosya: {os.path.basename(self.excel_path)}\n")
                f.write(f"Toplam Değişiklik: {len(self.change_log)}\n")
                f.write("=" * 80 + "\n\n")

                if not self.change_log:
                    f.write("Hiçbir değişiklik yapılmadı.\n")
                else:
                    for i, change in enumerate(self.change_log, 1):
                        f.write(f"DEĞİŞİKLİK #{i}\n")
                        f.write("-" * 80 + "\n")
                        f.write(f"Satır: {change['satir']}\n")
                        f.write(f"Sütun: {change['sutun']}\n")
                        f.write(f"Eski Değer: {change['eski_deger']}\n")
                        f.write(f"Yeni Değer: {change['yeni_deger']}\n")
                        f.write(f"Kural: {change['kural']}\n")
                        f.write("-" * 80 + "\n\n")

                f.write("=" * 80 + "\n")
                f.write("RAPOR SONU\n")
                f.write("=" * 80 + "\n")

        except Exception as e:
            print(f"Log dosyası kaydedilemedi: {e}")
            messagebox.showwarning("Uyarı", f"Değişiklik raporu kaydedilemedi:\n{str(e)}")

    def process_excel(self):
        if not self.excel_path:
            messagebox.showwarning("Uyarı", "Lütfen önce bir Excel dosyası seçin!")
            return

        try:
            self.status_var.set("Excel dosyası yükleniyor...")
            self.root.update()

            # Excel dosyasını yükle
            self.workbook = openpyxl.load_workbook(self.excel_path)
            self.worksheet = self.workbook.active

            # Log'u temizle
            self.change_log = []

            current_year = datetime.now().year
            changes_made = 0

            self.status_var.set("Kontroller yapılıyor...")
            self.root.update()

            # Satırları işle (başlık satırını atla)
            max_row = self.worksheet.max_row

            for row in range(2, max_row + 1):
                if row % 100 == 0:
                    self.status_var.set(f"İşleniyor... Satır: {row}/{max_row}")
                    self.root.update()

                # C sütunu değerini al
                c_value = self.get_cell_value(row, "C")
                c_num = self.to_number(c_value)

                # Kural 1: C=3 ve AD=1 → AD ve A hücrelerini kırmızı yap
                if c_num == 3:
                    ad_value = self.get_cell_value(row, "AD")
                    ad_num = self.to_number(ad_value)
                    if ad_num == 1:
                        if self.set_cell_red(row, "AD", "Kural 1: C=3 ve AD=1 → AD hücresi kırmızı"):
                            changes_made += 1
                        if self.set_cell_red(row, "A", "Kural 1: C=3 ve AD=1 → A hücresi kırmızı"):
                            changes_made += 1

                # Kural 11: C=1 veya C=2 ve W=2 → AC=1 ve AD=1
                if c_num in [1, 2]:
                    w_value = self.get_cell_value(row, "W")
                    w_num = self.to_number(w_value)
                    if w_num == 2:
                        if self.set_cell_value(row, "AC", 1, f"Kural 11: C={c_num} ve W=2 → AC=1"):
                            changes_made += 1
                        if self.set_cell_value(row, "AD", 1, f"Kural 11: C={c_num} ve W=2 → AD=1"):
                            changes_made += 1

                # Kural 2: AU > mevcut yıl → AU kırmızı
                au_value = self.get_cell_value(row, "AU")
                au_num = self.to_number(au_value)
                if au_num and au_num > current_year:
                    if self.set_cell_red(row, "AU", f"Kural 2: AU ({au_num}) > {current_year} → AU hücresi kırmızı"):
                        changes_made += 1

                # Kural 3: AU>2000 ve C=3 → AV=7
                if au_num and au_num > 2000 and c_num == 3:
                    if self.set_cell_value(row, "AV", 7, f"Kural 3: AU ({au_num}) > 2000 ve C=3 → AV=7"):
                        changes_made += 1

                # Kural 4: C=15 → U=3
                if c_num == 15:
                    if self.set_cell_value(row, "U", 3, "Kural 4: C=15 → U=3"):
                        changes_made += 1

                # CY değerini al (sonraki kurallar için)
                cy_value = self.get_cell_value(row, "CY")
                cy_num = self.to_number(cy_value)

                # Kural 5: C=3 ve CY<50 → BL=2
                if c_num == 3 and cy_num and cy_num < 50:
                    if self.set_cell_value(row, "BL", 2, f"Kural 5: C=3 ve CY ({cy_num}) < 50 → BL=2"):
                        changes_made += 1

                # Kural 6: C=3 ve CY<100 → BN=2
                if c_num == 3 and cy_num and cy_num < 100:
                    if self.set_cell_value(row, "BN", 2, f"Kural 6: C=3 ve CY ({cy_num}) < 100 → BN=2"):
                        changes_made += 1

                # Kural 7: C=3 ve CY<100 → BO=1
                if c_num == 3 and cy_num and cy_num < 100:
                    if self.set_cell_value(row, "BO", 1, f"Kural 7: C=3 ve CY ({cy_num}) < 100 → BO=1"):
                        changes_made += 1

                # Kural 8: BP max 3
                bp_value = self.get_cell_value(row, "BP")
                bp_num = self.to_number(bp_value)
                if bp_num and bp_num > 3:
                    if self.set_cell_value(row, "BP", 3, f"Kural 8: BP ({bp_num}) > 3 → BP=3"):
                        changes_made += 1

                # Kural 9: BU > S → BU=S
                bu_value = self.get_cell_value(row, "BU")
                s_value = self.get_cell_value(row, "S")
                bu_num = self.to_number(bu_value)
                s_num = self.to_number(s_value)

                if bu_num and s_num and bu_num > s_num:
                    if self.set_cell_value(row, "BU", s_num, f"Kural 9: BU ({bu_num}) > S ({s_num}) → BU=S ({s_num})"):
                        changes_made += 1

                # Kural 10: DB sütunu değerlerini virgülden sonra max 2 haneye yuvarla
                db_value = self.get_cell_value(row, "DB")
                db_num = self.to_number(db_value)
                if db_num is not None:
                    # 2 ondalık basamağa yuvarla
                    rounded_db = round(db_num, 2)
                    if db_num != rounded_db:
                        if self.set_cell_value(row, "DB", rounded_db, f"Kural 10: DB ({db_num}) → 2 ondalık hane ({rounded_db})"):
                            changes_made += 1

                # Kural 12: AU değeri 4 haneden fazla ya da az ise B ve AU hücrelerini yeşil yap
                au_value = self.get_cell_value(row, "AU")
                if au_value is not None:
                    au_str = str(au_value).strip()
                    # Ondalık nokta veya virgül varsa sadece tam sayı kısmını al
                    if '.' in au_str:
                        au_str = au_str.split('.')[0]
                    if ',' in au_str:
                        au_str = au_str.split(',')[0]

                    # 4 hane kontrolü (tam sayı uzunluğu)
                    if len(au_str) != 4:
                        if self.set_cell_green(row, "B", f"Kural 12: AU ({au_value}) 4 haneli değil → B hücresi yeşil"):
                            changes_made += 1
                        if self.set_cell_green(row, "AU", f"Kural 12: AU ({au_value}) 4 haneli değil → AU hücresi yeşil"):
                            changes_made += 1

                # Kural 13: C=3,6,15 ve AZ boş ise AZ=998
                if c_num in [3, 6, 15]:
                    az_value = self.get_cell_value(row, "AZ")
                    if az_value is None or (isinstance(az_value, str) and az_value.strip() == ""):
                        if self.set_cell_value(row, "AZ", 998, f"Kural 13: C={c_num} ve AZ boş → AZ=998"):
                            changes_made += 1

                # Kural 14: C=3 veya C=6 ise AW=1
                if c_num in [3, 6]:
                    if self.set_cell_value(row, "AW", 1, f"Kural 14: C={c_num} → AW=1"):
                        changes_made += 1

                # Kural 15: C=15 ise AW=4
                if c_num == 15:
                    if self.set_cell_value(row, "AW", 4, f"Kural 15: C=15 → AW=4"):
                        changes_made += 1

                # Kural 16: C=6 ve CS boş ise CS=2
                if c_num == 6:
                    cs_value = self.get_cell_value(row, "CS")
                    if cs_value is None or (isinstance(cs_value, str) and cs_value.strip() == ""):
                        if self.set_cell_value(row, "CS", 2, f"Kural 16: C=6 ve CS boş → CS=2"):
                            changes_made += 1

            # Dosyayı kaydet
            self.status_var.set("Değişiklikler kaydediliyor...")
            self.root.update()

            # Orijinal dosyanın yanına _düzeltilmiş ekleyerek kaydet
            base_name = os.path.splitext(self.excel_path)[0]
            ext = os.path.splitext(self.excel_path)[1]
            new_file_path = f"{base_name}_düzeltilmiş{ext}"

            self.workbook.save(new_file_path)

            # Değişiklik logunu kaydet
            self.status_var.set("Değişiklik raporu oluşturuluyor...")
            self.root.update()

            log_file_path = f"{base_name}_değişiklik_raporu.txt"
            self.save_change_log(log_file_path)

            self.status_var.set("İşlem tamamlandı!")
            messagebox.showinfo(
                "Başarılı",
                f"İşlem tamamlandı!\n\n"
                f"Toplam {changes_made} değişiklik yapıldı.\n"
                f"Yeni dosya: {os.path.basename(new_file_path)}\n"
                f"Değişiklik raporu: {os.path.basename(log_file_path)}"
            )

        except Exception as e:
            self.status_var.set("Hata oluştu!")
            messagebox.showerror("Hata", f"İşlem sırasında hata oluştu:\n{str(e)}")
        finally:
            if self.workbook:
                self.workbook.close()


def main():
    root = tk.Tk()
    app = ExcelKontrolProgram(root)
    root.mainloop()


if __name__ == "__main__":
    main()
